import serial
from threading import Event, Thread
import serial.tools.list_ports
import random
import sys
import time
import multiprocessing as mp
from collections import deque
from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series

global closed
closed = False
global target_choices
target_choices = deque([])

class data_object:
    def __init__(self, *args):
        self.args = args


class Stack:
    def __init__(self):
        self.items = []

    def isEmpty(self):
        return self.items == []

    def push(self, item):
        self.items.append(item)

    def pop(self):
        return self.items.pop()

    def peek(self):
        if not (self.isEmpty()):
            return self.items[len(self.items) - 1]
        else:
            return None

    def size(self):
        return len(self.items)


getMegaDataStack = Stack()
send_mega_stack = Stack()


def gpio_blinker(color, loop_count, Pi):
    if Pi:
        if loop_count % 2 == 0:
            color.on()
        else:
            color.off()


def loop_counter(loop_number):
    loop_number += 1
    if loop_number >= 10:
        loop_number = 1
    return loop_number

def findMEGA():
    ports = list(serial.tools.list_ports.comports())
    for p in ports:
        if "0042" in p[2]:
            return (p[0])
    return ('NULL')

def MegaData(MEGA, send_flag, send_data, send_mega_main, kill_event, close_event):

    mega_data = data_object
    close_event = close_event
    global closed
    global target_choices
    target_start = False
    update_list = False
    already_sent = False
    
    while not close_event.is_set() and not closed:
        getdata = False
        if MEGA.is_open:
            try:
                startMarker = MEGA.read().decode()
            except Exception as e:
                print('[MegaDataThread] : ERROR: ' + str(e))
            else:
                if startMarker == "<":
                    megaDataStr = MEGA.read(25)
                    startTime = time.time()
                    megaDataTemp = list(megaDataStr.decode())
                    megaDataTemp.insert(0, startMarker)
                    megaData = megaDataTemp[:megaDataTemp.index(">") + 1]
                    tempData = "".join(megaData)
                    if target_start:
                        update_list = True
                        target_start = False
                    getdata = True

                    if send_flag.is_set():
                        try:
                            data = send_data.pop()
                            if data == "<-1,-1,0,-1,0,0,0>" and not already_sent:
                                target_start = True
                                already_sent = True
                            MEGA.write(data.encode())
                            print("[MegaDataThread] : SENT MEGA DATA:  " + data)
                        except serial.SerialException as e:
                            print("[MegaDataThread] : ** MEGA SEND FAILED **   " + str(e))
                            while (time.time() - startTime >= 0.5):
                                try:
                                    MEGA.write(data.encode())
                                    print("[MegaDataThread] : Launch motors starting...")
                                except serial.SerialException as e:
                                    print("[MegaDataThread] : ** MEGA SEND FAILED **   " + str(e))
                                    if time.time() - startTime >= 0.5:
                                        pass
                                else:
                                    send_flag.clear()
                        except:
                            print("[MegaDataThread] : Issue sending data")
                            pass
                        else:
                            send_flag.clear()

            finally:
                if getdata == True:
                    try:
                        tempData = tempData.strip("<")
                        tempData = tempData.strip(">")
                        tempData = tempData.split(",")
                        if not update_list:
                            lidar_2_Distance = int(tempData[0])  # lidarDistance = int(cm)
                            temperature = int(tempData[1])  # temperature = int()
                            voiceCommand = int(tempData[2])  # voice commands = int(from 1 to 5)
                            targetTiming = int(tempData[3])  # targetTiming = float(0.0)
                            targetBallSpeed = float(tempData[4])  # targetBallSpeed

                            mega_data.lidar_2_Distance = lidar_2_Distance
                            mega_data.temperature = temperature
                            mega_data.voiceCommand = voiceCommand
                            mega_data.targetTiming = targetTiming
                            mega_data.targetBallSpeed = targetBallSpeed

                            # if targetTiming or targetBallSpeed != 0.0:
                            #     print("targetTiming: ", targetTiming, " ||  targetBallSpeed: ", targetBallSpeed)

                            # Send to Launcher Parent (STACK)
                            getMegaDataStack.push(mega_data)
                            # Send to Main_File Process (QUEUE)
                            send_mega_main.put(tempData)
                        else:
#                            print(tempData)
                            one = int(tempData[0])
#                            print("got one")
                            if one == 1:
                                target_choices.append(1)
                            two = int(tempData[1])
#                            print("got two")
                            if two == 1:
                                target_choices.append(2)
                            three = int(tempData[2])
#                            print("got three")
                            if three == 1:
                                target_choices.append(3)
                            four = int(tempData[3])
#                            print("got four")
                            if four == 1:
                                target_choices.append(4)
                            if len(target_choices) > 0:
                                update_list = False
#                            print("[MegaDataThread]: Selected Targets: ",target_choices," Number: ", len(target_choices))
                            
                    except IndexError as i:
                        pass
#                        print("[MegaDataThread]: ")
                    except Exception as e:
                        print("[MegaDataThread] : Parsing Data:  " + str(e))
        else:
            print("[MegaDataThread] : ** Mega didnt respond  **")
            time.sleep(0.01)
            continue


class Launcher(mp.Process):
    '''# _____________________________________LAUNCHER PROCESS __________________________________#
Connects and communicates with the Arduino Mega: Launcher Motors, Ball Feeder, Wifi, Accelerometer? (Rangefinder?)
Launcher(Thread): ---> Arduino MEGA provide Angle values to both motors, request temperature's from the Uno, and distance measurements
RECEIVE from MEGA:
- Lidar_2_Distance
- temperature
- voiceCommand
- targetTiming

SEND to MEGA:
- MotorSpeed1
- MotorSpeed2
- targetChoice
- difficulty
- estimated_tof
- targetBallSpeed

Stack IO
RECEIVE from STACKS:
- guiStack
- getfutureDist
- stereo_pitch_launch
- sendfinalDistStack


SEND to STACKS:
- sendMegaDataStack
- sendLidar2Stack
- sendfinalDistStack
- sendTemperatureStack'''

    def __init__(self, gui_data, send_mega_data, get_final_dist, get_future_dist, pi_no_pi, led_color, kill_event,
                 pause_event, py_reset, launch_event, voice_cont, drill_results, player_name, launcher_close, rec_launchdist, future_event,yaw_boost,running_player):
        super(Launcher, self).__init__()
        self.send_mega_data = send_mega_data
        self.gui_data = gui_data
        self.get_final_dist = get_final_dist
        self.get_future_dist = get_future_dist
        self.send_flag = Event()
        self.working_on_the_Pi = pi_no_pi
        self.kill_event = kill_event
        self.pause_event = pause_event
        self.py_reset = py_reset
        self.launch_event = launch_event
        self.voice_control = voice_cont
        self.drill_result = drill_results
        self.player = player_name
        self.launcher_shutdown = launcher_close
        self.rec_launch_flag = rec_launchdist
        self.future_flag = future_event
        self.dynamic_yaw = yaw_boost
        self.player_running = running_player
        if self.working_on_the_Pi:
            self.color = led_color

        self.drillSpeed = 1
        self.difficulty = 1
        self.drillType = ""
        self.drillCount = 1
        # VOICE COMMANDS
        self.footVC = 1
        self.chestVC = 2
        self.headVC = 3
        self.beginVC = 4
        self.stopVC = 5
        self.pauseVC = 6
        self.leftVC = 7
        self.rightVC = 8

        self.chest_flag = Event()
        self.head_flag = Event()
        self.left_curve_flag = Event()
        self.right_curve_flag = Event()

#        self.chest_flag.set()
        # self.head_flag.set()

        # self.MEGA = None
        self.OLD_FUT_FINAL_DIST = None
        self.FUT_FINAL_DIST = None
        self.LaunchTime = 0.0
        self.launch_loop_count = 1
        self.stereoData = None
        self.stereo_Distance = 0.0
        self.rationaleDistMeasures = 0
        self.distanceTotal = 0
        self.used_distance = 0.0
        self.drill_data = data_object
        self.targetTimes = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        self.ballSpeeds = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
        self.target_list = []
        self.initiation = False

        # LAUNCHER OPTIONS:
        self.WAIT_TIME = 0
        self.DYNAMIC_WAIT_TIME = 15  # seconds
        self.STATIC_WAIT_TIME = 15  # seconds
        self.ballfeed = "0"
        # MEGA_DATA Stack values:
        self.voiceCommand = -1  # < set to 1 for testing the launcher (-1 otherwise)
        self.targetTiming = 0
        self.targetBallSpeed = 0.0
        self.startTime = 0.0
        self.first_drill = True
        self.old_target_time = 0
        self.old_ballspeed = 0.0
        self.close_event = Event()
        self.last_target = 0
        self.targetChoice = 0

        self.MIN_PERIOD = 37000
        
    def drill_wait_time(self):
        if self.LaunchTime != 0.0:
            print("Entering Drill Waittime")
            while (time.time() - self.LaunchTime) < self.WAIT_TIME and not self.kill_event.is_set():
                if self.pause_event.is_set():
                    print("[Launcher] : Paused Drill")
                    while self.pause_event.is_set():
                        time.sleep(1)

                self.first_drill = False
                time.sleep(.1)

                if self.kill_event.is_set():
                    self.close_drill()

                    # sys.exit()
        elif self.drillCount == 1:
            print("[Launcher] : First Ball")
            self.first_drill = True

    def wait_for_voice(self):
        while not self.voiceCommand == self.beginVC and not self.kill_event.is_set():
            try:
                self.MEGA_DATA = getMegaDataStack.peek()
                self.voiceCommand = self.MEGA_DATA.voiceCommand
                self.temperature = self.MEGA_DATA.temperature
                if self.temperature != 0.0:
                    self.drill_data.temp = self.temperature
                if self.voiceCommand > 0:
                    self.handle_voicecommand()
                time.sleep(0.5)
            except:
                print("[Launcher] : Waiting for Voice Command")
                time.sleep(1)
                continue

    def handle_voicecommand(self):
        if self.voiceCommand == self.stopVC:
            print("[VOICE_COMMAND]: STOP received")
            self.kill_event.set()
        elif self.voiceCommand == self.footVC:
            print("[VOICE_COMMAND]: footVC received")
            if self.chest_flag.is_set():
                self.chest_flag.clear()
            elif self.head_flag.is_set():
                self.head_flag.clear()
            else:
                print("[VOICE_COMMAND]: Already set to foot")
        elif self.voiceCommand == self.chestVC:
            print("[VOICE_COMMAND]: chestVC received")
            if not self.chest_flag.is_set():
                self.chest_flag.set()
            else:
                print("[VOICE_COMMAND]: Already set to chest")
        elif self.voiceCommand == self.headVC:
            print("[VOICE_COMMAND]: headVC received")
            if not self.head_flag.is_set():
                self.head_flag.set()
            else:
                print("[VOICE_COMMAND]: Already set to head")
        elif self.voiceCommand == self.beginVC:
            self.pause_event.clear()
        elif self.voiceCommand == self.pauseVC:
            print("[VOICE_COMMAND]: PAUSE received")
            if not self.pause_event.is_set:
                self.pause_event.set()
            else:
                print("[VOICE_COMMAND]: already Paused")
        else:
            print("[VOICE_COMMAND]: UNIDENTIFIED: ", self.voiceCommand)

    def wait_for_target(self):
        target_data = False
        start = time.time()
        self.get_mega_data()
        while not target_data and not self.kill_event.is_set():
            try:
                i = self.drillCount - 2
                if self.targetTiming != 0 and self.targetTiming != self.old_target_time:
                    if i != 0:
                        print("[Launcher] : targetTiming = " + str(self.targetTiming), "    ||", i)
                    self.targetTimes[i] = self.targetTiming
                    self.old_target_time = self.targetTiming
                    target_data = True

                if self.targetBallSpeed != 0.0 and self.targetBallSpeed != self.old_ballspeed:
                    if i != 0:
                        print("[Launcher] : targetBallSpeed = " + str(self.targetBallSpeed), " ||", i)
                    self.ballSpeeds[i] = self.targetBallSpeed
                    self.old_ballspeed = self.targetBallSpeed
                    target_data = True

                if self.pause_event.is_set():
                    print("[Launcher] : Paused Drill")
                    while self.pause_event.is_set():
                        time.sleep(1)

                end = time.time() - start

                if end > 12:
                    target_data = True
            except Exception as e:
                print("[Launcher]: ->", e)

        if self.kill_event.is_set():
            self.close_drill()

    def get_mega_data(self):
        got_mega = False
        while not got_mega:
            try:
                self.MEGA_DATA = getMegaDataStack.peek()
                self.voiceCommand = self.MEGA_DATA.voiceCommand
                self.temperature = self.MEGA_DATA.temperature
                self.targetTiming = self.MEGA_DATA.targetTiming
                self.targetBallSpeed = self.MEGA_DATA.targetBallSpeed

            except:
                print("[Launcher] : Waiting for MEGA DATA")
                time.sleep(1)
                continue
            else:
                if self.temperature != 0.0:
                    self.drill_data.temp = self.temperature
                # time.sleep(0.5)
                if self.voiceCommand > 0:
                    self.handle_voicecommand()
                got_mega = True

    def send_launch_data(self):
        global target_choices
        # ____POLYNOMIAL FIT FROM THEORETICAL VALUES___ #
#        if self.drillType == "Dynamic":
#                self.dynamic_yaw.set()
                
        if self.used_distance == 0.0:
            self.used_distance = 0.1
        elif self.used_distance == None:
            self.used_distance = 0.1
            
        PERIOD = int(117952.849005218 * self.used_distance ** -0.481698152749178 + 12254.5437)

        if self.chest_flag.is_set():    ## NEEDS SCALING
            PERIOD = PERIOD + (37276*0.17)

            if PERIOD < self.MIN_PERIOD:
                PERIOD = self.MIN_PERIOD

            PERIOD1 = PERIOD
            PERIOD2 = PERIOD
            
        elif self.head_flag.is_set():   ## NEEDS SCALING
            PERIOD = PERIOD + (37276*0.2)

            if PERIOD < self.MIN_PERIOD:
                PERIOD = self.MIN_PERIOD

            PERIOD1 = PERIOD
            PERIOD2 = PERIOD
        else:
            if PERIOD < self.MIN_PERIOD:
                PERIOD = self.MIN_PERIOD

            PERIOD1 = PERIOD
            PERIOD2 = PERIOD

        motor_period1 = str(int(PERIOD1))
        motor_period2 = str(int(PERIOD2))

#        print("[Launcher]: ____ ", target_choices)
        if len(target_choices) > 0 and len(self.target_list) == 0:
            for i in target_choices:
                self.target_list.append(i)
            if len(self.target_list) == 0:
                self.target_list = [0]
                
            # print("[Launcher]: TARGET LIST = ", self.target_list)
            
        if len(self.target_list) > 0:
            self.targetChoice = int(random.choice(self.target_list))
            
        estimated_tof = (0.120617 * self.used_distance) * 1000
        estimated_tof = round(estimated_tof, 2)

        if self.first_drill and not self.initiation:
            time.sleep(5)
            print("Sending Target Initiation:")

            self.feed = "0"
            self.diff = "-1"
            drill_type = "0"
            estimated_tof = 0    
            targetChoice = "0"
            motor_period1 = "-1"
            motor_period2 = "-1"
            self.send_data = '<' + motor_period1 + ',' + motor_period2 + ',' + str(self.targetChoice) + ',' + str(
                self.diff) + ',' + self.feed + ',' + str(estimated_tof) + ',' + drill_type + '>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            while self.send_flag.is_set():
                time.sleep(0.1)
            self.initiation = True

            time.sleep(5)
        else:
            if self.drillType == "Dynamic":
                drill_type = "1"
#                self.dynamic_yaw.set()  # <----!!!
            else:
                drill_type = "0"

            self.send_data = '<' + motor_period1 + ',' + motor_period2 + ',' + str(self.targetChoice) + ',' + str(
                self.difficulty) + ',' + self.ballfeed + ',' + str(estimated_tof) + ',' + drill_type + '>'

            if self.drillType == "Dynamic":
                self.dynamic_yaw.set()  # <----!!!

            # ____________________ Write data to MEGA ____________________
            send_mega_stack.push(self.send_data)
            print("[Launcher] : BALL NUMBER :  ", self.drillCount)

            # if self.dynamic_yaw.is_set():
            #     time.sleep(1)
            # self.launch_event.set()
            self.send_flag.set()
#            if self.drillType == "Dynamic":
#                self.dynamic_yaw.set()  # <----!!!
            while self.send_flag.is_set():
                time.sleep(0.1)
                    
            if self.ballfeed == "1":
                # self.launch_event.set()
                print("[Launcher] : LAUNCHING BALL")
                self.drillCount += 1  # << ON SUCCESSFUL LAUNCH
                self.LaunchTime = time.time()
#            if self.drillCount >= 5:
#                self.wait_for_target()
#                self.close_drill()

    def launcher_startup(self):
        startData = False
        while not startData and not self.kill_event.is_set() and not self.pause_event.is_set():
            self.drillSpeed = self.gui_data.speed
            self.difficulty = self.gui_data.difficulty
            self.drillType = self.gui_data.drilltype

            if self.pause_event.is_set():
                print("[Launcher] : Paused Drill")
                while self.pause_event.is_set():
                    time.sleep(1)

            try:
                mega_port = findMEGA()
                MEGA = serial.Serial(mega_port, 115200, timeout=1)
                MEGA.baudrate = 115200

                # START ARDUINO COMMUNICATIONS THREAD
                self.mega_data_thread = Thread(target=MegaData,
                                               args=[MEGA, self.send_flag, send_mega_stack, self.send_mega_data,
                                                     self.kill_event, self.close_event])
                self.mega_data_thread.start()
            except serial.SerialException as err:
                print("[Launcher] : MEGA thread failed to start" + str(err))
                time.sleep(2)
                continue
            except Exception as nodata:
                print("[Launcher] : ** Mega data thread failed to start:  " + str(nodata) + "**")
                time.sleep(2)
                # Get TEMPERATURE and WAIT for voiceCommand
            else:
                print("[Launcher] : Connected to MEGA")
                # _____________________WAIT FOR VOICE COMMAND and TEMP TO BEGIN DRILL _____________________
                if self.voice_control.is_set():
                    self.wait_for_voice()
                else:
                    time.sleep(3)
                # ****SET TEMPERATURE CORRECTION FACTOR*****
                self.tempCorrect = 1  # self.temperature / 25
                # *****************************************
                startData = True

    def launcher_common(self):
        self.startTime = time.time()

        gpio_blinker(self.color, self.launch_loop_count, self.working_on_the_Pi)

        self.launch_loop_count = loop_counter(self.launch_loop_count)

        if not self.first_launch:
            self.wait_for_target()
        else:
            self.first_launch = False
            self.ballfeed = "0"
            time.sleep(3)
            self.send_launch_data()

        self.drill_wait_time()

        if self.pause_event.is_set():
            print("[Launcher] : Paused Drill")
            while self.pause_event.is_set():
                time.sleep(1)

    def get_distance(self):
        launch_data = False
        while not launch_data and not self.kill_event.is_set():
            try:
                self.rec_launch_flag.set()
                FINAL_DIST = self.get_final_dist.get(timeout=1)  # <<<<<
                self.rec_launch_flag.clear()
#                print("[Launcher] :  Final Dist : " + str(FINAL_DIST))
            except:
                print("[Launcher] : Waiting for final distance")
                time.sleep(2)
                continue
            else:
                self.used_distance = FINAL_DIST
                launch_data = True
        if self.kill_event.is_set():
#            self.wait_for_target()
            self.close_drill()
            
    def dynamic_drill(self):
        self.WAIT_TIME = self.DYNAMIC_WAIT_TIME
        print("[Launcher] : Beginning Drill")
        self.first_launch = True
        self.launcher_common()

        while self.drillCount <= 5 and not self.kill_event.is_set():
            if self.pause_event.is_set():
                print("[Launcher] : Paused Drill")
                while self.pause_event.is_set():
                    time.sleep(1)

            self.get_mega_data()

#            if self.player_running.is_set() and time.time() - self.LaunchTime > 7:
#                self.player_running.clear()
            warmup = time.time()
            while time.time() - warmup < 3: # and not future_dist:
                self.get_distance()
                self.ballfeed = "0"
                self.send_launch_data()
                self.get_mega_data()
            self.dynamic_yaw.set()
#            time.sleep(0.4)
            self.future_flag.set()
#            self.dynamic_yaw.set()
            
#            warmup = time.time()
#            while time.time() - warmup < 0.1: # and not future_dist:
#                self.get_distance()
#                self.ballfeed = "0"
#                self.send_launch_data()
#                self.get_mega_data()
            
            try:
                self.rec_launch_flag.set() #< !!!
                self.FUT_FINAL_DIST = self.get_future_dist.get(timeout=1)  # <<<<< GET PREDICTED LOCATION
                self.rec_launch_flag.clear() #< !!!

                print("[Launcher]: FUT_FINAL_DIST:  ", self.FUT_FINAL_DIST)
            except TimeoutError:
                print("[Launcher]: getting Future distance timed out")
                self.future_flag.clear()
                self.get_distance()
                self.ballfeed = "1"
#                self.dynamic_yaw.set()  # <----!!!
                self.send_launch_data()
#                    self.dynamic_yaw.set()  # <----!!!
            except:
                print("[Launcher]failed to get future dist")
                self.future_flag.clear()
                self.get_distance()
#                self.used_distance = self.FUT_FINAL_DIST
                self.ballfeed = "1"
#                self.dynamic_yaw.set()  # <----!!!
                self.send_launch_data()
#                    self.dynamic_yaw.set()  # <----!!!

            else:
                self.future_flag.clear()
                self.used_distance = self.FUT_FINAL_DIST
                self.ballfeed = "1"
#                self.dynamic_yaw.set()  # <----!!!
                self.send_launch_data()
#                    self.dynamic_yaw.set()  # <----!!!

            self.player_running.clear()
            self.launcher_common()


        if self.drillCount >= 5:
            self.wait_for_target()
        self.close_drill()

    def static_drill(self):
        self.WAIT_TIME = self.STATIC_WAIT_TIME
        print("[Launcher] : Beginning Drill")
        self.first_launch = True

        while self.drillCount <= 5 and not self.kill_event.is_set():
            self.launcher_common()
            self.get_mega_data()

        # SEND LOOP TO WARMuP MOTORS
            warmup = time.time()
            while time.time() - warmup < 4:
                self.get_distance()
                self.ballfeed = "0"
                self.send_launch_data()
                self.get_mega_data()
                if self.kill_event.is_set():
                    self.close_drill()
            if self.kill_event.is_set():
                self.close_drill()
            self.get_distance()
            self.ballfeed = "1"
            self.send_launch_data()
            self.get_mega_data()

        if self.drillCount >= 5:
            self.wait_for_target()
        self.close_drill()

    def manual_drill(self):
        self.WAIT_TIME = 0  # <<< WAIT TIME IS HANDLED DIFFERENTLY HERE
        print("[Launcher] : Beginning Drill")
        self.first_launch = True

        while self.drillCount <= 5 and not self.kill_event.is_set():
            # WAIT FOR VC FOR EVERY LAUNCh
            if not self.first_launch:
                self.wait_for_voice()

            self.launcher_common()
            self.get_mega_data()

            # SEND LOOP TO WARMuP MOTORS
            warmup = time.time()
            while time.time() - warmup < 4:
                self.get_distance()
                self.ballfeed = "0"
                self.send_launch_data()
                self.get_mega_data()
                if self.kill_event.is_set():
                    self.close_drill()
            if self.kill_event.is_set():
                self.close_drill()
            self.get_distance()
            self.ballfeed = "1"
            self.send_launch_data()
            self.get_mega_data()

        if self.drillCount >= 5:
            self.wait_for_target()
        self.close_drill()

    def output_to_excel(self):
        filename = "/home/pi/Desktop/demo_wb.xlsx"

        try:
            wb = load_workbook(filename=filename)
            print("[Launcher]: file exists")
        except Exception as e:
            print("[Launcher]: file not found, making new book ")
            wb = Workbook()
            summary_sheet = wb.get_sheet_by_name('Sheet')
            summary_sheet.title = 'Summary'

        try:
            players_sheet = wb[self.player]
            print("[Launcher]: sheet exists, appending data")

        except KeyError as k:
            print("[Launcher]: creating new sheet for: ", self.player)
            players_sheet = wb.create_sheet(title=self.player)

        data_printed = False
        col = 1
        row = 1
        drill = 1
        while not data_printed:
            if players_sheet.cell(column=col, row=row + 1).value is None:
                players_sheet.cell(column=col, row=row + 1).value = "Drill" + str(drill)
                if row == 1:
                    players_sheet.cell(column=col + 1, row=row).value = "Ball #"
                    players_sheet.cell(column=col + 2, row=row).value = "Time for Pass"
                    players_sheet.cell(column=col + 3, row=row).value = "Ball Speed"
                for i in range(5):
                    players_sheet.cell(column=col + 1, row=row + i + 1).value = row + i
                    players_sheet.cell(column=col + 2, row=row + i + 1).value = (self.targetTimes[i + 1] / 1000)
                    players_sheet.cell(column=col + 3, row=row + i + 1).value = self.ballSpeeds[i + 1]

                data_printed = True

            else:
                row += 5
                drill += 1

        wb.save(filename)
        # PLOT THE DATA:
        ranges = [0] * (len(wb.worksheets) - 1)
        speed_ranges = [0] * (len(wb.worksheets) - 1)
        series = [0] * 20

        chart1 = ScatterChart()
        chart2 = ScatterChart()
        chart3 = ScatterChart()
        chart4 = ScatterChart()
        chart5 = ScatterChart()
        charts = [chart1, chart2, chart3, chart4, chart5]

        speed_chart1 = ScatterChart()
        speed_chart2 = ScatterChart()
        speed_chart3 = ScatterChart()
        speed_chart4 = ScatterChart()
        speed_chart5 = ScatterChart()
        speed_charts = [speed_chart1, speed_chart2, speed_chart3, speed_chart4, speed_chart5]

        for i in range((len(wb.worksheets) - 1)):
            #            print(str(wb.worksheets[i+1])[11:20])
            gathered = False
            coll = 1
            roww = 1
            speed_charts[i].title = str(wb.worksheets[i + 1])[12:19] + "'s  Results"
            charts[i].title = str(wb.worksheets[i + 1])[12:19] + "'s  Results"

            while not gathered:
                print("Gathering..")
                if wb.worksheets[i + 1].cell(column=coll, row=roww + 1).value is None:
                    gathered = True
                else:
                    roww += 5
                    gathered = False

            ranges[i] = Reference(wb.worksheets[i + 1], min_col=3, min_row=1, max_col=3, max_row=roww)
            x_vals = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)
            series[i] = Series(ranges[i], xvalues=x_vals, title_from_data=True)
            charts[i].append(series[0])

            charts[i].x_axis.title = "Test Number"
            charts[i].x_axis.scaling.min = 0
            charts[i].x_axis.scaling.max = roww
            charts[i].y_axis.title = "Reaction Times (seconds)"
            charts[i].legend = None

            s1 = charts[i].series[0]
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
            s1.graphicalProperties.line.noFill = True

            speed_ranges[i] = Reference(wb.worksheets[i + 1], min_col=4, min_row=1, max_col=4, max_row=roww)
            x_value = Reference(wb.worksheets[i + 1], min_col=2, min_row=1, max_row=roww)
            series[i + 10] = Series(speed_ranges[i], xvalues=x_value, title_from_data=True)
            speed_charts[i].append(series[i + 10])

            speed_charts[i].x_axis.title = "Test Number"
            speed_charts[i].x_axis.scaling.min = 0
            speed_charts[i].x_axis.scaling.max = roww
            speed_charts[i].y_axis.title = "Ball Speed (m/s)"
            speed_charts[i].legend = None

            s2 = speed_charts[i].series[0]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "006868"  # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "006868"  # Marker outline
            s2.graphicalProperties.line.noFill = True

            charts[i].style = 13
            wb.worksheets[0].add_chart(charts[i], "A" + str(i + 1 + 15 * i))
            speed_charts[i].style = 13
            wb.worksheets[0].add_chart(speed_charts[i], "J" + str(i + 1 + 15 * i))

        wb.save(filename)

    def display_results(self):
        results = [0, 0]
        results[0] = self.targetTimes
        results[1] = self.ballSpeeds
        self.drill_result.put(results)

    def close_drill(self):
        print("[Launcher] : Drill Ending  ")
        global closed
#        self.launcher_shutdown.set()
        if self.kill_event.is_set():
            print("[Launcher] :  Closing process")
            self.send_data = '<-1,-1,0,6,0,0.0>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            start = time.time()
            while self.send_flag.is_set() or self.py_reset.is_set:
                time.sleep(0.1)
#                if self.py_reset.is_set:
#                    print("waiting for py_reset")
#                elif self.send_flag.is_set:
#                    print("waiting for send_flag reset")
                if time.time() - start > 4:
                    print("[Launcher]: Closing Drill")
                    print("==================================")
                    print("        Drill Closed")
                    print("==================================")                    
                    break
                    
            closed = True
            self.close_event.set()
            sys.exit()

        elif self.drillCount >= 5:
            last_data = False
            start = time.time()
            i = self.drillCount - 2

            while not last_data:
                try:
                    i = 5
                    self.MEGA_DATA = getMegaDataStack.peek()
                    self.voiceCommand = self.MEGA_DATA.voiceCommand
                    self.temperature = self.MEGA_DATA.temperature
                    self.targetTiming = self.MEGA_DATA.targetTiming
                    self.targetBallSpeed = self.MEGA_DATA.targetBallSpeed

                    if self.temperature != 0.0:
                        #                        print("VC = " + str(self.voiceCommand) + "  TEMP = " + str(self.temperature))
                        self.drill_data.temp = self.temperature
                    #                    else:
                    #                        print("VC = ", str(self.voiceCommand))

                    time.sleep(0.5)
                except:
                    print("[Launcher] : Waiting for Voice Command")
                    time.sleep(1)
                    continue
                else:
                    try:
                        print("Waiting for ball 5 target data")
                        if self.targetTiming != 0 and self.targetTiming != self.old_target_time:
                            self.targetTimes[i] = self.targetTiming
                            self.old_target_time = self.targetTiming
                            # if self.targetTimes[4] != self.targetTimes[5]:
                            print("[Launcher] : targetTiming = ", str(self.targetTiming), " |last| ", i)
                            last_data = True

                        if self.targetBallSpeed != 0.0 and self.targetBallSpeed != self.old_ballspeed:
                            self.ballSpeeds[i] = self.targetBallSpeed
                            self.old_ballspeed = self.targetBallSpeed
                            # if self.ballSpeeds[4] != self.ballSpeeds[5]:
                            print("[Launcher] : targetBallSpeed = ", str(self.targetBallSpeed), " |last| ", i)
                            last_data = True

                        end = time.time() - start

                        if end > 12:
                            last_data = True
                    except Exception as e:
                        print("[Launcher]: close_drill error", e)

            time.sleep(5)
            self.send_data = '<-1,-1,0,6,0,0.0>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            while self.send_flag.is_set():
                time.sleep(0.1)
            self.close_event.set()
            closed = True
            self.mega_data_thread.join()

            print("[Launcher] : Drill COMPLETE!")
            print("======================================================================")
            print(self.player, "'s results for ", self.drillType, " passing drill:")
            print("User selcted a Difficulty: ", self.difficulty, " and a Speed: ", self.drillSpeed)
            print()
            print("Ball #               :          one    |    two    |     three    |    four    |    five    ")
            try:
                print("Player reaction time :      ", self.targetTimes[1], "  ", self.targetTimes[2], "  ",
                      self.targetTimes[3], "  ", self.targetTimes[4], "  ", self.targetTimes[5], "  ")
                print("Player passing speed :      ", self.ballSpeeds[1], "  ", self.ballSpeeds[2], "  ",
                      self.ballSpeeds[3], "  ", self.ballSpeeds[4], "  ", self.ballSpeeds[5], "  ")
                print("======================================================================")


            except AttributeError as a:
                pass
            try:
                print("At a current temperature of :        ", self.drill_data.temp)
            except:
                pass

            self.output_to_excel()  # NEW LINEs *!*!*!*!*!*!**!*!*!*!*!*!*!
            self.display_results()

            self.kill_event.set()
            sys.exit()

        else:
            print("[Launcher] : Not sure what went wrong")
            self.send_data = '<-1,-1,0,6,0,0.0>'
            send_mega_stack.push(self.send_data)
            self.send_flag.set()
            while self.send_flag.is_set() or self.py_reset.is_set:
                time.sleep(0.1)
            self.launcher_shutdown.clear()
            sys.exit()

    def run(self):
        print("[Launcher] : Starting ...")
        self.launcher_startup()

        try:
            if self.drillType == 'Dynamic':
                # ____________________________ DYNAMIC DRILL _____________________________ #
                self.dynamic_drill()

            if self.drillType == 'Static':
                # ____________________________ STATIC DRILL _____________________________ #
                self.static_drill()

            if self.drillType == 'Manual':
                # ____________________________ MANUAL DRILL _____________________________ #
                self.manual_drill()

        except Exception as e:
            print('[Launcher] : failed because of exception: ' + str(e))
            self.kill_event.set()
            self.close_drill()